import click

from openpyxl import load_workbook
from ac_api import Description, IndicatorType, Indicator, IndicatorGroup, \
    IdString, Template, Model, PrimaryTemplate, Equipment
from mapping import *
from typing import List

@click.group()
def cli():
    """ Root for the CLI """
    pass


@cli.command()
@click.argument("datafile", type=click.Path(exists=True))
def load(datafile):
    """ Load AC data from a spreadsheet
    Inserts all of the given data into AC.
    Writes the AC ids back into spreadsheet.

    Args:
        datafile - xlsx file that contains the data to be loaded
    """
    click.echo("Opening %s..." % datafile)
    wb = load_workbook(filename=datafile)
    indicators = load_indicators(wb["Indicator"])
    update_worksheet(indicators, wb["Indicator"])
    indicator_groups = load_indicator_groups(indicators, wb["Indicator Group"])
    update_worksheet(indicator_groups, wb["Indicator Group"])
    templates = load_templates(indicator_groups, wb["Model Template"])
    update_worksheet(templates, wb["Model Template"])
    models = load_models(templates, wb["Model"])
    update_worksheet(models, wb["Model"])
    equipment = load_equipment(models, wb["Equipment"])
    update_worksheet(equipment, wb["Equipment"])
    # save the changes
    wb.save(filename=datafile)

@cli.command()
@click.argument("datafile", type=click.Path(exists=True))
def delete(datafile):
    """ Delete AC data defined in spreadsheet
    Requires ids in the first column of each object to be deleted

    Args:
        datafile - the xlsx file that contains the data to be deleted
     """

    print(f"Opening {datafile}...")
    wb = load_workbook(filename=datafile)
    # do this in the reverse order of the loads due to dependencies
    equip_sheet = wb["Equipment"]
    for row in equip_sheet.iter_rows(min_row=2):
        if row[ID].value:
            equip = Equipment(equipmentId=row[ID].value)
            status = equip.delete()
            if status == 204:
                row[ID].value = ""
                print(f"Deleted {row[INTERNAL_ID].value}")
            else:
                print(f"Could not delete {row[ID].value}")

    model_sheet = wb["Model"]
    for row in model_sheet.iter_rows(min_row=2):
        if row[ID].value:
            model = Model(modelId=row[ID].value)
            status = model.delete()
            if status == 204:
                row[ID].value = ""
                print(f"Deleted {row[INTERNAL_ID].value}")
            else:
                print(f"Could not delete {row[ID].value}")

    template_sheet = wb["Model Template"]
    for row in template_sheet.iter_rows(min_row=2):
        if row[ID].value:
            template = Template(id=row[ID].value)
            status = template.delete()
            if status == 200:
                row[ID].value = ""
                print(f"Deleted {row[INTERNAL_ID].value}")
            else:
                print(f"Could not delete {row[ID].value}")

    indicator_group_sheet = wb["Indicator Group"]
    previous_id = ""
    for iteration, row in enumerate(indicator_group_sheet.iter_rows(min_row=2)):
        # check for duplicate IDs in the column and clear them
        if previous_id == row[ID].value:
            row[ID].value = ""

        if row[ID].value:
            indicator_group = IndicatorGroup(id=row[ID].value)
            status = indicator_group.delete()
            if status == 200:
                previous_id = row[ID].value
                row[ID].value = ""
                print(f"Deleted {row[INTERNAL_ID].value}")
            else:
                print(f"Could not delete {row[ID].value}")

    indicator_sheet = wb["Indicator"]
    for row in indicator_sheet.iter_rows(min_row=2):
        if row[ID].value:
            indicator = Indicator(id=row[ID].value)
            status = indicator.delete()
            if status == 200:
                row[ID].value = ""
                print(f"Deleted {row[INTERNAL_ID].value}")
            else:
                print(f"Could not delete {row[ID].value}")


    # save the changes
    wb.save(filename=datafile)



def load_indicators(indicator_sheet):
    """ Loads all of the indicators into AC

    Args:
        indicator_sheet - worksheet containing the required datafields

    Returns:
        List of indicators that were loaded

    """
    indicators = []

    # open the indicator sheet and load the objects
    for row in indicator_sheet.iter_rows(min_row=2, values_only=True):
        indicator = Indicator(internalId=row[IND_INTERNAL_ID],
                                description=Description(row[IND_DESCRIPTION]),
                                dataType=row[IND_DATA_TYPE],
                                dimension1=row[IND_DIMENSION],
                                indicatorUom=row[IND_UOM],
                                expectedBehaviour=str(row[IND_EXPECTED_BEHAVIOR]),
                                indicatorColorCode=row[IND_COLOR],
                                id=4)

        indicators.append(indicator)

    # insert into AC
    for indicator in indicators:
        print(f"inserting indicator {indicator.internalId}...")
        try:
            indicator.insert()
        except Exception as ex:
            print(f"failed...error:{ex}")
        else:
            print(f"sucess...id = {indicator.id}")

    return indicators

def load_indicator_groups(indicators: List[Indicator], ig_sheet):
    """ Loads all of the indicator groups into AC

    Args:
        indicators - list of indicators that were loaded
        ig_sheet - worksheet containing the required datafields


    Returns:
        List of indicator groups that were loaded
    """

    # open the indicator group sheet and load the objects
    # loop through the row and get the distinct indicator group identifiers
    indicator_groups = []
    internal_id = ""
    desc = ""
    ig_indicators = []
    for iteration, row in enumerate(ig_sheet.iter_rows(min_row=2, values_only=True)):
        # first iteration
        if iteration == 0:
            internal_id = row[IG_INTERNAL_ID]
            desc = row[IG_DESCRIPTION]
            ig_indicators.append(row[IG_INDICATOR])
        else:
            # new internal id?
            if internal_id != row[IG_INTERNAL_ID]:
                # create indicator group and add to list
                indicator_groups.append(IndicatorGroup(internalId=internal_id,
                    description=Description(desc), indicators=ig_indicators))
                ig_indicators = []
                internal_id = row[IG_INTERNAL_ID]
                desc = row[IG_DESCRIPTION]
                ig_indicators.append(row[IG_INDICATOR])
            else: # same internal id, append the indicator
                ig_indicators.append(row[IG_INDICATOR])

    # out of the loop, we should have at least one indicator group
    indicator_groups.append(IndicatorGroup(internalId=internal_id,
        description=Description(desc), indicators=ig_indicators))

    # get the ids for each indicator in each indicator group
    for indicator_group in indicator_groups:
        # loop through the temp_ids in the indicator group
        for iteration, temp_id in enumerate(indicator_group.indicators):
            for ind in indicators:
                # get the real id from the indicators list and replace the temp_id
                if ind.internalId == temp_id:
                    indicator_group.indicators[iteration] = ind.id
                    break

    # insert into AC
    for indicator_group in indicator_groups:
        print(f"inserting indicator group {indicator_group.internalId}...")
        try:
            indicator_group.insert()
        except Exception as ex:
            print(f"failed...error: {ex}")
        else:
            print(f"success...id = {indicator_group.id}")

    return indicator_groups

def load_templates(indicator_groups, template_sheet):
    """ Loads all of the templates into AC

    Supports model templates only.

    Args:
        indicator_groups - list of indicator groups that were loaded
        template_sheet - worksheet that contains required datafields

    Returns:
        list of the templates that were loaded

    """
    # open the template sheet and load the objects
    # loop through the row and get the distinct template identifiers
    templates = []
    internal_id = ""
    desc = ""
    template_ig = []
    for iteration, row in enumerate(template_sheet.iter_rows(min_row=2, values_only=True)):
        # first iteration
        if iteration == 0:
            internal_id = row[TEM_INTERNAL_ID]
            desc = row[TEM_DESCRIPTION]
            template_ig.append(row[TEM_INDICATOR_GROUP])
        else:
            # new internal id?
            if internal_id != row[TEM_INTERNAL_ID]:
                # create template and add to list
                templates.append(Template(internalId=internal_id,
                    description=Description(desc), indicatorGroups=template_ig))
                template_ig = []
                internal_id = row[TEM_INTERNAL_ID]
                desc = row[TEM_DESCRIPTION]
                template_ig.append(row[TEM_INDICATOR_GROUP])
            else: # same internal id, append the indicator group
                template_ig.append(row[TEM_INDICATOR_GROUP])

    # out of the loop, we should have at least one indicator group
    templates.append(Template(internalId=internal_id,
        description=Description(desc), indicatorGroups=template_ig))

    # get the ids for each indicator group in each template
    for template in templates:
        # loop through the temp_ids in the template
        for iteration, temp_id in enumerate(template.indicatorGroups):
            for ig in indicator_groups:
                # get the real id from the indicator group list and replace the temp_id
                if ig.internalId == temp_id:
                    template.indicatorGroups[iteration] = IdString(ig.id)
                    break

    # insert into AC
    for template in templates:
        print(f"inserting template {template.internalId}...")
        try:
            template.insert()
        except Exception as ex:
            print(f"failed...error: {ex}")
        else:
            print(f"success...id = {template.id}")

    return templates

def load_models(templates, model_sheet):
    """ Loads all of the models into AC

    Args:
        templates - list of templates that were loaded
        model_sheet - worksheet that contains required datafields

    Returns:
        list of the models that were loaded

    """
    # open the model sheet and load the objects
    # loop through the row and get the distinct identifiers
    models = []
    for row in model_sheet.iter_rows(min_row=2, values_only=True):
       # create model and add to list
       models.append(Model(internalId=row[MOD_INTERNAL_ID],
            description=row[MOD_DESCRIPTION],
            templates=row[MOD_TEMPLATE],
            equipmentTracking=row[MOD_TRACKING],
            organizationID=row[MOD_ORG]))

    # get the ids for the template in each row
    for model in models:
        for template in templates:
            # get the real id from the indicator group list and replace the temp_id
            if model.templates == template.internalId:
                model.templates = [PrimaryTemplate(template.id)]
                break

    # insert into AC
    for model in models:
        print(f"inserting and publishing model {model.internalId}...")
        try:
            model.insert()
            model.publish()
        except Exception as ex:
            print(f"failed...error: {ex}")
        else:
            print(f"success...id = {model.modelId}")

    return models


def load_equipment(models, equipment_sheet):
    equipment_list = []
    for row in equipment_sheet.iter_rows(min_row=2, values_only=True):
        equipment_list.append(Equipment(internalId=row[EQU_INTERNAL_ID],
            description=Description(row[EQU_DESCRIPTION]),
            modelId=row[EQU_MODEL],
            operatorID=row[EQU_OPERATOR],
            lifeCycle=row[EQU_LIFECYCLE]))

    for equipment in equipment_list:
        for model in models:
            if equipment.modelId == model.internalId:
                equipment.modelId = model.modelId
                break

    for equipment in equipment_list:
        print(f"inserting {equipment.internalId}...")
        try:
            equipment.insert()
        except Exception as ex:
            print(f"failed...error: {ex}")
        else:
            print(f"success...id = {equipment.equipmentId}")

    return equipment_list



def update_worksheet(asset_central_objects: List, worksheet):
    """ Update the worksheet with returned IDs in the first column

    Args:
        asset_central_objects - a list of objects with ids
        worksheet - the worksheet to be updated
    """
    # check that we have a list to delete
    if asset_central_objects is None:
        return
    # get the internal ids from the spreadsheet
    for iteration, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
        internal_id = row[INTERNAL_ID]
        for obj in asset_central_objects:
            if obj.internalId == internal_id:
                # check for model since it has a different id
                if isinstance(obj, Model):
                    value = obj.modelId
                elif isinstance(obj, Equipment):
                    value = obj.equipmentId
                else:
                    value = obj.id
                # adjust cells to account for header row and 1-based counting
                worksheet.cell(column=ID+1, row=iteration+2, value=value)
                break


